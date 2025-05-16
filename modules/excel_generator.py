import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel_files(parsed_data, report_type, year, months, temp_dir):
    """
    Создает Excel файлы с данными
    
    Args:
        parsed_data (dict): Обработанные данные партнеров
        report_type (str): Тип отчета ('free' или 'paid')
        year (str): Год данных
        months (list): Список месяцев
        temp_dir (str): Директория для сохранения файлов
    
    Returns:
        str: Путь к созданному Excel файлу
    """
    # Определяем тип отчета и настраиваем название файла
    if report_type == 'free':
        file_title = f"Льготные_1C_ИТС_{year}"
        data_key = 'free_data'
        title = "Льготные 1C:ИТС"
    else:  # 'paid'
        file_title = f"Платные_1C_ИТС_{year}"
        data_key = 'paid_data'
        title = "Платные 1C:ИТС"
    
    # Диагностика
    print(f"Создание Excel-файла: {title}")
    print(f"Количество партнеров: {len(parsed_data['partners'])}")
    print(f"Количество месяцев: {len(parsed_data['months'])}")
    
    # Если нет данных или месяцев, создаем пустой файл
    if not parsed_data['months'] or not parsed_data['partners']:
        print(f"Недостаточно данных для создания отчета")
        empty_path = os.path.join(temp_dir, f"{file_title}.xlsx")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет"
        ws["A1"] = "Нет данных для отчета"
        wb.save(empty_path)
        
        return empty_path
    
    # Создаем DataFrame
    # Сначала подготовим список всех месяцев для колонок
    sorted_months = sorted(parsed_data['months'], key=lambda x: (x['year'], x['month']))
    
    # Отладочная информация о месяцах
    for month_info in sorted_months:
        print(f"Месяц в отчете: {month_info['month']:02d}.{month_info['year']}")
    
    # Создаем список колонок
    columns = ["Номер", "Название", "Город"]
    
    # Добавляем колонки месяцев и разницы
    month_columns = []
    for i, month_info in enumerate(sorted_months):
        month_name = f"{month_info['month']:02d}.{month_info['year']}"
        month_columns.append(month_name)
        columns.append(month_name)
        
        # Добавляем колонку для разницы (кроме первого месяца)
        if i > 0:
            prev_month_info = sorted_months[i-1]
            prev_month_name = f"{prev_month_info['month']:02d}.{prev_month_info['year']}"
            diff_col = f"Разница {prev_month_name}-{month_name}"
            columns.append(diff_col)
    
    # Создаем пустой DataFrame
    df = pd.DataFrame(columns=columns)
    
    # Заполняем данными
    rows = []
    for idx, (name, partner_data) in enumerate(parsed_data['partners'].items(), 1):
        row = {
            "Номер": idx,
            "Название": name,
            "Город": partner_data['city']
        }
        
        # Добавляем данные по месяцам
        for i, month_info in enumerate(sorted_months):
            month_key = month_info['key']
            month_name = f"{month_info['month']:02d}.{month_info['year']}"
            
            # Значение за месяц
            month_value = partner_data[data_key].get(month_key, 0)
            row[month_name] = month_value
            
            # Разница с предыдущим месяцем
            if i > 0:
                prev_month_info = sorted_months[i-1]
                prev_month_key = prev_month_info['key']
                prev_month_name = f"{prev_month_info['month']:02d}.{prev_month_info['year']}"
                
                prev_value = partner_data[data_key].get(prev_month_key, 0)
                diff = month_value - prev_value
                
                diff_col = f"Разница {prev_month_name}-{month_name}"
                row[diff_col] = diff
        
        rows.append(row)
    
    # Создаем DataFrame из списка строк
    df = pd.DataFrame(rows, columns=columns)
    
    # Сортируем по последнему месяцу (если есть месяцы)
    if month_columns:
        last_month = month_columns[-1]
        df = df.sort_values(by=last_month, ascending=False).reset_index(drop=True)
        
        # Обновляем номера после сортировки
        df['Номер'] = range(1, len(df) + 1)
    
    # Диагностика DataFrame
    print(f"Создан DataFrame размером {df.shape}")
    print(f"Колонки DataFrame: {df.columns.tolist()}")
    
    # Путь для сохранения
    file_path = os.path.join(temp_dir, f"{file_title}.xlsx")
    
    # Сохраняем в Excel
    writer = pd.ExcelWriter(file_path, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='Отчет')
    
    # Форматируем Excel
    workbook = writer.book
    worksheet = writer.sheets['Отчет']
    
    # Форматирование заголовка
    for col_num, column_title in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Цвет фона заголовков
        if 'Разница' in column_title:
            cell.fill = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')  # Светло-зеленый
        else:
            cell.fill = PatternFill(start_color='C9DAF8', end_color='C9DAF8', fill_type='solid')  # Светло-голубой
    
    # Автоподбор ширины колонок
    for col_num, column in enumerate(df.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_num)
        
        # Учитываем заголовок
        max_length = max(max_length, len(str(column)))
        
        # Перебираем все ячейки в колонке
        for row_num in range(2, len(df) + 2):
            cell = worksheet.cell(row=row_num, column=col_num)
            
            # Форматирование числовых ячеек
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='center')
                
                # Особое форматирование разницы
                if 'Разница' in column:
                    if cell.value > 0:
                        cell.font = Font(color='006100')  # Зеленый для положительных
                    elif cell.value < 0:
                        cell.font = Font(color='9C0006')  # Красный для отрицательных
            
            # Учитываем длину ячейки
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        # Устанавливаем ширину колонки
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем форматированный файл
    writer.close()
    
    print(f"Excel-файл успешно создан: {file_path}")
    return file_path